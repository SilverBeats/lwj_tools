#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import csv
import json
import pickle
from typing import Any, Iterator, List, Union

import numpy as np
import pandas as pd
import yaml
from openpyxl.reader.excel import load_workbook

from .tools import ext_check
from ..utils.common import get_file_name_and_ext


class FileReader:
    @staticmethod
    @ext_check(ext=["json"])
    def read_json(file_path: str, **json_kwargs) -> Any:
        """读取json文件

        Args:
            file_path: json 文件路径
            **json_kwargs:

        Returns:
            Any: json文件内容
        """
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f, **json_kwargs)

    @staticmethod
    @ext_check(ext=["jsonl"])
    def read_jsonl(file_path: str, return_iter: bool = False, **json_kwargs) -> Union[List[Any], Iterator[Any]]:
        """读取jsonl文件

        Args:
            file_path: jsonl 文件路径
            return_iter: 是否返回迭代器
            **json_kwargs: json.loads的参数

        Returns:
            List[Any] | Iterator[Any]: `return_iter` 为 False 时返回列表，为 True 时返回迭代器
        """

        def parse_fn():
            with open(file_path, "r", encoding="utf-8") as json_file:
                for line in json_file:
                    yield json.loads(line, **json_kwargs)

        if return_iter:
            return parse_fn()
        else:
            return list(parse_fn())

    @staticmethod
    @ext_check(ext=["txt"])
    def read_txt(file_path: str, return_iter: bool = False) -> Union[List[Any], Iterator[Any]]:
        """读取txt文件

        Args:
            file_path: txt 文件路径
            return_iter: 是否返回迭代器

        Returns:
            List[str] | Iterator[str]: `return_iter` 为 False 时返回列表，为 True 时返回生成器
        """

        def parse_fn():
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    yield line.strip()

        if return_iter:
            return parse_fn()
        else:
            return list(parse_fn())

    @staticmethod
    @ext_check(ext=["yaml", "yml"])
    def read_yaml(file_path) -> dict:
        """读取yaml文件

        Args:
            file_path: yaml 文件路径

        Returns:
            dict: yaml文件内容
        """
        with open(file_path, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f)
        return config

    @staticmethod
    @ext_check(ext=["pkl"])
    def read_pkl(file_path: str, **pickle_kwargs) -> Any:
        """读取pkl文件

        Args:
            file_path: pkl 文件路径
            **pickle_kwargs: `pickle.load`的参数

         Returns:
            Any: pkl 文件内容
        """
        with open(file_path, "rb") as f:
            return pickle.load(f, **pickle_kwargs)

    @staticmethod
    @ext_check(ext=["npy", "npz"])
    def read_npyz(file_path, **numpy_kwargs):
        """读取npy或npz文件

        Args:
            file_path: npy或npz文件路径
            **numpy_kwargs: `numpy.load`的参数

        Returns:
            Any: npy或npz文件内容
        """
        return np.load(file_path, **numpy_kwargs)

    @staticmethod
    @ext_check(ext=["xlsx", "xls"])
    def read_large_excel(file_path: str, sheet_name: str = "Sheet1") -> Iterator[dict]:
        wb = load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]
        _iter = ws.iter_rows()
        header = [cell.value for cell in next(_iter)]
        for row in _iter:
            data = dict(zip(header, (cell.value for cell in row)))
            yield data

    @staticmethod
    @ext_check(ext=["xlsx", "xls"])
    def read_excel(
        file_path: str,
        sheet_name: str = "Sheet1",
        return_iter: bool = False,
        return_dict: bool = False,
    ) -> Union[pd.DataFrame, Iterator[dict], List[dict]]:
        """读取excel文件

        Args:
            file_path: excel 文件路径
            sheet_name: sheet 名称
            return_iter: 是否返回生成器
            return_dict: 是否返回字典

        Returns:
            pd.DataFrame | Iterator[dict] | List[dict]: `return_iter` 为 False 时返回 DataFrame，为 True
            时返回生成器，return_dict 为 True 时返回字典列表
        """
        if return_iter:
            return_dict = True

        if return_iter:
            return FileReader.read_large_excel(file_path, sheet_name)
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if return_dict:
            return df.to_dict("records")
        return df

    @staticmethod
    @ext_check(ext=["csv", "tsv"])
    def read_csv(
        file_path: str,
        delimiter: str = ",",
        return_iter: bool = False,
        return_dict: bool = False,
        has_header: bool = True,
        **pd_kwargs,
    ) -> Union[pd.DataFrame, Iterator[dict], List[dict]]:
        """读取csv文件

        Args:
            file_path: csv 文件路径
            delimiter: 分隔符
            return_iter: 是否返回迭代器
            return_dict: 是否返回字典
            has_header: 是否有表头
            **pd_kwargs: `pandas.read_csv`的参数

        Returns:
            pd.DataFrame | Iterator[dict] | List[dict]: `return_iter` 为 False 时返回 DataFrame，为 True 时返回迭代器，return_dict
            为 True 时返回字典列表
        """

        def parse_fn():
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f, delimiter=delimiter)
                column_names = None
                if has_header and return_dict:
                    column_names = [s.strip() for s in next(reader)]
                elif return_dict:
                    first_row = next(reader)
                    column_names = list(range(len(first_row)))
                    yield dict(zip(column_names, first_row))

                for row in reader:
                    if not return_dict:
                        yield row
                    else:
                        yield dict(zip(column_names, row))

        if get_file_name_and_ext(file_path, False)[-1] == "tsv":
            delimiter = "\t"

        if return_iter:
            return parse_fn()
        else:
            df = pd.read_csv(
                file_path,
                delimiter=delimiter,
                encoding="utf-8",
                **pd_kwargs,
            )
            if return_dict:
                return df.to_dict("records")
            return df

    @staticmethod
    @ext_check(
        ext=[
            "json", "jsonl", "txt", "yaml", "yml", "pkl",
            "xlsx", "xls", "csv", "tsv", "npy", "npz",
        ],
    )
    def read(
        file_path: str,
        return_iter: bool = False,
        return_dict: bool = False,
        delimiter: str = ",",
        has_header: bool = True,
        sheet_name: str = "Sheet1",
        **specific_kwargs,
    ):
        """读取文件

        Args:
            file_path: 文件路径
            return_iter: 是否返回迭代器
            return_dict: 是否返回字典
            delimiter: 分隔符
            has_header: 是否有表头
            sheet_name: sheet 名称
            **specific_kwargs: 特定文件的参数

        Returns:
            Any: 文件内容
        """
        ext = get_file_name_and_ext(file_path, False)[-1]

        kwargs = {}
        if ext in ["jsonl", "txt", "xlsx", "xls", "csv", "tsv"]:
            kwargs["return_iter"] = return_iter

        if ext in ["xlsx", "xls", "csv", "tsv"]:
            kwargs["return_dict"] = return_dict

        if ext in ["xlsx", "xls"]:
            kwargs["sheet_name"] = sheet_name

        if ext in ["csv", "tsv"]:
            kwargs["delimiter"] = delimiter
            kwargs["has_header"] = has_header

        if ext == "json":
            func = FileReader.read_json
        elif ext == "jsonl":
            func = FileReader.read_jsonl
        elif ext == "txt":
            func = FileReader.read_txt
        elif ext in ["yaml", "yml"]:
            func = FileReader.read_yaml
        elif ext == "pkl":
            func = FileReader.read_pkl
        elif ext in ["xlsx", "xlx"]:
            func = FileReader.read_excel
        elif ext in ["csv", "tsv"]:
            func = FileReader.read_csv
        else:  # ext in ["npy", "npz"]
            func = FileReader.read_npyz

        return func(file_path, **kwargs, **specific_kwargs)
