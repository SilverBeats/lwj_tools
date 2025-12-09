#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import csv
import json
import os
import pickle
from typing import Any, Callable, Iterable, List, Union

import numpy as np
import pandas as pd
from omegaconf import DictConfig, OmegaConf
from openpyxl.reader.excel import load_workbook

from .tools import get_file_name_and_ext
from ..errors import FileReadError, FileTypeError


def ext_check(ext: Union[str, List[str]]):
    """
    Args:
        ext: one or multi file extensions without dot.
    """

    def decorator(func: Callable):
        def wrapper(*args, **kwargs):
            file_path = kwargs.get("file_path", None)
            if file_path is None and len(args) > 0:
                func_name = func.__name__
                file_path_index = 0 if "read" in func_name else 1
                file_path = args[file_path_index]

            if not file_path:
                raise FileReadError(f"{file_path} is required")

            file_ext = get_file_name_and_ext(file_path, with_dot=False)[-1]

            allowed_exts = [ext] if isinstance(ext, str) else ext
            if file_ext not in allowed_exts:
                allowed_str = ", ".join(allowed_exts)
                raise FileTypeError(f"{file_path} is not a {allowed_str} file!")

            return func(*args, **kwargs)

        return wrapper

    return decorator


class FileReader:
    @staticmethod
    @ext_check(
        ext=[
            "json",
            "jsonl",
            "txt",
            "yaml",
            "yml",
            "pkl",
            "xlsx",
            "xls",
            "csv",
            "tsv",
            "npy",
            "npz",
        ],
    )
    def read(
        file_path: str,
        return_iter: bool = False,
        return_dict: bool = False,
        delimiter: str = ",",
        has_header: bool = True,
        sheet_name: str = "Sheet1",
        **specific_kwargs,
    ):
        ext = get_file_name_and_ext(file_path, False)[-1]

        kwargs = {}
        if ext in ["jsonl", "txt", "xlsx", "xls", "csv", "tsv"]:
            kwargs["return_iter"] = return_iter

        if ext in ["yaml", "yml", "xlsx", "xls", "csv", "tsv"]:
            kwargs["return_dict"] = return_dict

        if ext in ["xlsx", "xls"]:
            kwargs["sheet_name"] = sheet_name

        if ext in ["csv", "tsv"]:
            kwargs["delimiter"] = delimiter
            kwargs["has_header"] = has_header

        if ext == "json":
            func = FileReader.read_json
        elif ext == "jsonl":
            func = FileReader.read_jsonl
        elif ext == "txt":
            func = FileReader.read_txt
        elif ext in ["yaml", "yml"]:
            func = FileReader.read_config
        elif ext == "pkl":
            func = FileReader.read_pkl
        elif ext in ["xlsx", "xlx"]:
            func = FileReader.read_excel
        elif ext in ["csv", "tsv"]:
            func = FileReader.read_csv
        else:  # ext in ["npy", "npz"]
            func = FileReader.read_npyz

        return func(file_path, **kwargs, **specific_kwargs)

    @staticmethod
    @ext_check(ext=["json"])
    def read_json(file_path: str, **json_kwargs):
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f, **json_kwargs)

    @staticmethod
    @ext_check(ext=["jsonl"])
    def read_jsonl(file_path: str, return_iter: bool = False, **json_kwargs):

        def parse_fn():
            with open(file_path, "r", encoding="utf-8") as json_file:
                for line in json_file:
                    yield json.loads(line, **json_kwargs)

        if return_iter:
            return parse_fn()
        else:
            return list(parse_fn())

    @staticmethod
    @ext_check(ext=["txt"])
    def read_txt(file_path: str, return_iter: bool = False):

        def parse_fn():
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    yield line.strip()

        if return_iter:
            return parse_fn()
        else:
            return list(parse_fn())

    @staticmethod
    @ext_check(ext=["yaml", "yml"])
    def read_config(file_path, return_dict: bool = False):
        with open(file_path, "r", encoding="utf-8") as f:
            config = OmegaConf.load(f)
        if return_dict:
            return OmegaConf.to_container(config, resolve=True, enum_to_str=True)
        return config

    @staticmethod
    @ext_check(ext=["pkl"])
    def read_pkl(file_path: str, **pickle_kwargs):
        with open(file_path, "rb") as f:
            return pickle.load(f, **pickle_kwargs)

    @staticmethod
    @ext_check(ext=["xlsx", "xls"])
    def read_large_excel(file_path: str, sheet_name: str = "Sheet1") -> Iterable[dict]:
        wb = load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]
        _iter = ws.iter_rows()
        header = [cell.value for cell in next(_iter)]
        for row in _iter:
            data = dict(zip(header, (cell.value for cell in row)))
            yield data

    @staticmethod
    @ext_check(ext=["xlsx", "xls"])
    def read_excel(
        file_path: str,
        sheet_name: str = "Sheet1",
        return_iter: bool = False,
        return_dict: bool = False,
    ) -> Union[pd.DataFrame, Iterable[dict], List[dict]]:
        if return_iter:
            return_dict = True

        if return_iter:
            return FileReader.read_large_excel(file_path, sheet_name)
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if return_dict:
            return df.to_dict("records")
        return df

    @staticmethod
    @ext_check(ext=["csv", "tsv"])
    def read_csv(
        file_path: str,
        delimiter: str = ",",
        return_iter: bool = False,
        return_dict: bool = False,
        has_header: bool = True,
        **pd_kwargs,
    ):
        def parse_fn():
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f, delimiter=delimiter)
                column_names = None
                if has_header and return_dict:
                    column_names = [s.strip() for s in next(reader)]
                elif return_dict:
                    # Dummy column names if no header but return dict
                    first_row = next(reader)
                    column_names = list(range(len(first_row)))
                    yield dict(zip(column_names, first_row))

                for row in reader:
                    if not return_dict:
                        yield row
                    else:
                        yield dict(zip(column_names, row))

        if get_file_name_and_ext(file_path, False)[-1] == "tsv":
            delimiter = "\t"

        if return_iter:
            return parse_fn()
        else:
            df = pd.read_csv(
                file_path,
                delimiter=delimiter,
                encoding="utf-8",
                **pd_kwargs,
            )
            if return_dict:
                return df.to_dict("records")
            return df

    @staticmethod
    @ext_check(ext=["npy", "npz"])
    def read_npyz(file_path, **numpy_kwargs):
        return np.load(file_path, **numpy_kwargs)


class FileWriter:

    @staticmethod
    @ext_check(
        ext=[
            "json",
            "jsonl",
            "txt",
            "yaml",
            "yml",
            "pkl",
            "xlsx",
            "xls",
            "csv",
            "tsv",
            "npy",
            "npz",
        ],
    )
    def dump(data: Any, file_path: str, sheets: str = "Sheet1", **specific_kwargs):
        ext = get_file_name_and_ext(file_path, False)[-1]
        kwargs = {}
        if ext in ["xlsx", "xls"]:
            kwargs["sheets"] = sheets

        if ext == "json":
            func = FileWriter.dump_json
        elif ext == "jsonl":
            func = FileWriter.dump_jsonl
        elif ext == "txt":
            func = FileWriter.dump_txt
        elif ext in ["yaml", "yml"]:
            func = FileWriter.dump_config
        elif ext == "pkl":
            func = FileWriter.dump_pkl
        elif ext in ["xlsx", "xlx"]:
            func = FileWriter.dump_excel
        elif ext in ["csv", "tsv"]:
            func = FileWriter.dump_csv
        else:  # ext in ["npy", "npz"]
            func = FileWriter.dump_npyz

        func(data, file_path, **kwargs, **specific_kwargs)

    @staticmethod
    @ext_check(ext=["json"])
    def dump_json(data: Any, file_path: str, **json_kwargs):
        os.makedirs(os.path.dirname(file_path) or "./", exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, **json_kwargs)

    @staticmethod
    @ext_check(ext=["jsonl"])
    def dump_jsonl(data: List[Any], file_path: str, **json_kwargs):
        os.makedirs(os.path.dirname(file_path) or "./", exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as json_file:
            for item in data:
                line = json.dumps(item, ensure_ascii=False, **json_kwargs)
                json_file.write(line + "\n")

    @staticmethod
    @ext_check(ext=["txt"])
    def dump_txt(data: List[str], file_path: str):
        os.makedirs(os.path.dirname(file_path) or "./", exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as txt_file:
            for line in data:
                txt_file.write(line.strip() + "\n")

    @staticmethod
    @ext_check(ext=["yaml", "yml"])
    def dump_config(data: Union[dict, DictConfig], file_path: str):
        os.makedirs(os.path.dirname(file_path) or "./", exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as f:
            OmegaConf.save(data, f)

    @staticmethod
    @ext_check(ext=["pkl"])
    def dump_pkl(data: Any, file_path: str, **pickle_kwargs):
        os.makedirs(os.path.dirname(file_path) or "./", exist_ok=True)
        with open(file_path, "wb") as f:
            pickle.dump(data, f, **pickle_kwargs)

    @staticmethod
    @ext_check(ext=["xlsx", "xls"])
    def dump_excel(
        data: Union[pd.DataFrame, List[pd.DataFrame]],
        file_path: str,
        sheets: Union[str, List[str]] = "Sheet1",
        **pd_kwargs,
    ):
        if isinstance(data, pd.DataFrame):
            data = [data]

        if isinstance(sheets, str):
            sheets = [sheets]

        assert len(data) == len(sheets), "data and sheets must have same length"
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            for df, sheet in zip(data, sheets):
                df.to_excel(writer, sheet_name=sheet, **pd_kwargs)

    @staticmethod
    @ext_check(ext=["csv", "tsv"])
    def dump_csv(data: pd.DataFrame, file_path: str, **pd_kwargs):
        os.makedirs(os.path.dirname(file_path) or "./", exist_ok=True)
        data.to_csv(file_path, **pd_kwargs)

    @staticmethod
    @ext_check(ext=["npy", "npz"])
    def dump_npyz(data: Any, file_path: str):
        os.makedirs(os.path.dirname(file_path) or "./", exist_ok=True)
        ext = get_file_name_and_ext(file_path, with_dot=False)[1]
        if ext == "npz":
            np.savez(file_path, data)
        else:
            np.save(file_path, data)
